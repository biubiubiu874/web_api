import openpyxl

class Excel:
    # 初始化excel操作对象时，先定位到具体的表单
    def __init__(self,file_path,num_sheet):
        self.file_path=file_path
        self.excel= openpyxl.load_workbook(r'{}'.format(self.file_path))
        self.sheet = self.excel[self.excel.sheetnames[num_sheet]]
    # 读取excel表格中有那些表单
    def read_sheet_name(self):
        print(self.excel.sheetnames)
    #读取所有数据
    def read_cell_values(self):
        total_data = list(self.sheet.rows)
        row=1
        title=[]
        dict_row_data=[]
        for i in total_data:
            row_data=[]
            # 将获取标题跟内容放在一起了，当row=1时候，获取标题，大于1时，获取内容，最后zip拼接在一起，dict_row_data就是测试数据
            for j in i:
                if row==1:
                    title.append(j.value)
                if row!=1:
                    row_data.append(j.value)
            if row!=1:
                dict_row_data.append(dict(zip(title,row_data)))
            row += 1
        return dict_row_data
    # 传入行列，和值，修改单元格
    def change_cell_value(self,row,col,value):
        self.sheet.cell(row,col).value=value
        self.excel.save(r'{}'.format(self.file_path))
    # 析构函数，最后关闭文件
    # def __del__(self):
        # self.excel.close()
# excel=Excel(r'D:\1234.xlsx')
# excel.read_sheet_name()
# dict_row_data=excel.read_cell_values()
# print(dict_row_data)